import csv
from openpyxl import Workbook

new_workbook = Workbook()

filename = './assests/maintained/data/district.csv'
maxPopulationDID = []
minPopulationDID = []

# Refer to ./refining_dataScripts/max_min_csv.py for the following snippet
max_population = [228848, 285387, 323870, 387570, 1204953]
low_population = [42821, 45714, 51313, 51428, 53921]

with open(filename, 'r', newline='', encoding='utf-8') as csvfile:
    csv_reader = csv.reader(csvfile)

    for row in csv_reader:
        if (row[3] == 'A4'):
            pass
        else:
            if (int(row[3]) in max_population):
                maxPopulationDID.append(row[0])
            elif (int(row[3]) in low_population):
                minPopulationDID.append(row[0])

filename_newAccount = './assests/maintained/data/new_account.csv'
minStage1OBJ = {}
maxStage1OBJ = {}
minStage1List = []
maxStage1List = []

with open(filename_newAccount, 'r', newline='', encoding='utf-8') as csvfile:
    csv_reader = csv.reader(csvfile)

    for row in csv_reader:
        if (row[1] in maxPopulationDID):
            maxStage1OBJ = {
                'account_id': row[0],
                'district_id': row[1],
            }
            maxStage1List.append(maxStage1OBJ)
        elif (row[1] in minPopulationDID):
            minStage1OBJ = {
                'account_id': row[0],
                'district_id': row[1],
            }
            minStage1List.append(minStage1OBJ)

# Considering last three months: ['MAY', 'JUN', 'JUL']

filename_newTransaction = './assests/maintained/data/new_transaction.csv'
month = [
    'MAY', 'JUN', 'JUL'
]
f1OBJ = []
temp = {}

with open(filename_newTransaction, 'r', newline='', encoding='utf-8') as csvfile:
    csv_reader = csv.reader(csvfile)

    for row in csv_reader:
        monthNum = str(row[2])[2:4]
        if (monthNum == '05'):
            monthName = 'MAY'
        elif (monthNum == '06'):
            monthName = 'JUN'
        elif (monthNum == '07'):
            monthName = 'JUL'
        else:
            monthName = 'N/A'

        if (monthName in month):
            temp = {
                'account_id': row[1],
                'month': monthName,
                'amount': row[5]
            }
            f1OBJ.append(temp)

# For 5 highly populated

max_populatedSheet = []
waste = []

for a in range(len(f1OBJ)):
    for b in range(len(maxStage1List)):
        if (f1OBJ[a]['account_id'] == maxStage1List[b]['account_id']):
            waste = [
                int(f1OBJ[a]['account_id']),
                f1OBJ[a]['month'],
                float(f1OBJ[a]['amount']),

            ]
            max_populatedSheet.append(waste)
            waste = []

max_sheet = new_workbook.create_sheet(title='Max Populated')

max_sheet.append(['Account ID', 'Month', 'Amount'])

for row_data in max_populatedSheet:
    max_sheet.append(row_data)

# For 5 lowest populated

min_populatedSheet = []
waste = []

for a in range(len(f1OBJ)):
    for b in range(len(minStage1List)):
        if (f1OBJ[a]['account_id'] == minStage1List[b]['account_id']):
            waste = [
                int(f1OBJ[a]['account_id']),
                f1OBJ[a]['month'],
                float(f1OBJ[a]['amount']),

            ]
            min_populatedSheet.append(waste)
            waste = []

min_sheet = new_workbook.create_sheet(title='Min Populated')

min_sheet.append(['Account ID', 'Month', 'Amount'])

for row_data in min_populatedSheet:
    min_sheet.append(row_data)

new_workbook.save('./assests/maintained/scripted_data/prob_state2.xlsx')

print('Done!')
